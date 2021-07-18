import datetime
import numpy as np
import os
from functools import partial
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import List, Iterable

from . import constants as C
from .session import SessionInfo
from ..tracker.info import TracksCompleteInfo, Direction
from ..tracker.track import Track

def EN(track: Track, info: TracksCompleteInfo):
    return info.track_direction(track) == Direction.EN

def SN(track: Track, info: TracksCompleteInfo):
    return info.track_direction(track) == Direction.SN

def loaded(tracks: Iterable[Track]):
    yield from (track for track in tracks if track.load_detected and track.load_prediction)

def unloaded(tracks: Iterable[Track]):
    yield from (track for track in tracks if track.load_detected and not track.load_prediction)

def onedim_conversion(length_or_speed: float, mm_per_pixel: float):
    """
    :param length_or_speed: in [px] or [px/s]
    :param mm_per_pixel: in [mm/px]
    :return: length i [mm] or speed in [mm/s]
    """
    return length_or_speed * mm_per_pixel

def area_conversion(area: float, mm_per_pixel: float):
    """
    :param area: in [px^2]
    :param mm_per_pixel: in [mm/px]
    :return: area in [mm^2]
    """
    return area * (mm_per_pixel ** 2)

def length(cell):
    if not cell.value: return 0
    if isinstance(cell.value, float):
        lc = len(str(round(cell.value, 2))) + 1
    elif isinstance(cell.value, datetime.datetime):
        lc = 19
    else:
        lc = len(str(cell.value)) + 1
    return lc

def adjust_column_widths(ws: Worksheet):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.row == 1: continue
            if cell.value:
                dims[cell.column_letter] = max(dims.get(cell.column_letter, 13), length(cell))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    from openpyxl.utils import get_column_letter
    for rng in ws.merged_cells.ranges:
        cols = {}
        for cell in rng.cells:
            col = get_column_letter(cell[1])
            cols[col] = max(ws.column_dimensions[col].width, length(ws.cell(cell[1], cell[0])))
        total_length = sum(cols.values())
        for col in cols:
            ws.column_dimensions[col].width = total_length / len(cols)

def center_headers(ws: Worksheet, rows=1):
    for row in ws.iter_rows(1, rows):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

def adjust_decimal_places(ws: Worksheet, decimal_places=2):
    for row in ws.rows:
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0' + ('.' + '0' * decimal_places) if decimal_places else ''

class Exporter:
    def __init__(self):
        self.__wb = None

    @property
    def workbook(self) -> Workbook:
        if self.__wb is None:
            raise AttributeError("Primero debe llamar a export(_progress)")
        return self.__wb

    def save(self, path: Path):
        self.workbook.save(path)

    def export(self, infos: List[TracksCompleteInfo], time_delta=datetime.timedelta(minutes=1)):
        list(self.export_progress(infos, time_delta))
        return self.__wb

    def export_progress(self, session: SessionInfo, time_delta=datetime.timedelta(minutes=1)):
        yield "Inicializando...", 0, 1
        wb = Workbook()

        # region Sheets y headers

        ant_sheet = wb.active
        ant_sheet.title = "Por hormiga"
        ant_sheet.append([
            "Video",
            "ID",
            "Dirección",
            "Carga",
            "Certeza en carga",
            "Velocidad prom. [mm/s]",
            "Área mediana [mm²]",
            "Largo mediana [mm]",
            "Ancho mediana [mm]",
            "Tiempo entrada",
            "Tiempo salida",
            "Frame inicio",
            "Frame final",
        ])
        video_sheet = wb.create_sheet("Por video")
        video_sheet.append(["Video",
                   "Hora Inicio",
                   "Hora Fin",
                   "Total hormigas entrando al nido (EN)", "",
                   "Total hormigas saliendo del nido (SN)", "",
                   "Velocidad promedio EN [mm/s]", "",
                   "Velocidad promedio SN [mm/s]", "",
                   "Área mediana EN [mm²]", "",
                   "Área mediana SN [mm²]", "",
                   "Largo mediana EN [mm]", "",
                   "Largo mediana SN [mm]", "",
                   "Ancho mediana EN [mm]", "",
                   "Ancho mediana SN [mm]", "",
                   ])
        video_sheet.append(["", "", ""] + ["Cargadas", "Sin carga"] * 10)
        merge = ['A1:A2', 'B1:B2', 'C1:C2', 'D1:E1', 'F1:G1', 'H1:I1', 'J1:K1', 'L1:M1', 'N1:O1', 'P1:Q1', 'R1:S1',
                 'T1:U1', 'V1:W1']
        for m in merge: video_sheet.merge_cells(m)
        props = ('speed_mean', 'area_median', 'length_median', 'width_median')

        time_sheet = wb.create_sheet("En el tiempo")
        time_sheet.append([
            "Hora Inicio",
            "Hora Fin",
            "Total hormigas entrando al nido (EN)", "",
            "Total hormigas saliendo del nido (SN)", "",
            "Velocidad promedio EN [mm/s]", "",
            "Velocidad promedio SN [mm/s]", "",
            "Área mediana EN [mm²]", "",
            "Área mediana SN [mm²]", "",
            "Largo mediana EN [mm]", "",
            "Largo mediana SN [mm]", "",
            "Ancho mediana EN [mm]", "",
            "Ancho mediana SN [mm]", "",
        ])
        time_sheet.append(["", ""] + ["Cargadas", "Sin carga"] * 10)
        merge = ['A1:A2', 'B1:B2', 'C1:D1', 'E1:F1', 'G1:H1', 'I1:J1', 'K1:L1', 'M1:N1', 'O1:P1', 'Q1:R1', 'S1:T1',
                 'U1:V1']
        for m in merge: time_sheet.merge_cells(m)

        # endregion

        yield "Inicializando...", 1, 1
        yield "Cargando información...", 0, 1

        trkfiles = [session.get_trkfile(f) for f in session.videofiles]
        start_times = {}
        end_times = {}
        for itrk, trkfile in enumerate(trkfiles):
            info: TracksCompleteInfo = TracksCompleteInfo.load(trkfile)
            progress_i = 0
            od = partial(onedim_conversion, mm_per_pixel=info.mm_per_pixel)
            area = partial(area_conversion, mm_per_pixel=info.mm_per_pixel)
            tracks = info.filter_tracks(**C.TRACKFILTER)
            progress_max = len(tracks) + 1
            # region Por hormiga

            for track in tracks:
                direction = info.track_direction(track)
                if direction != Direction.UN:
                    ant_sheet.append([
                        info.video_name,
                        track.id,
                        direction.name,
                        ("Si" if track.load_prediction else "No") if track.load_detected else "??",
                        track.load_certainty if track.load_detected else 0,
                        od(track.speed_mean),
                        area(track.area_median),
                        od(track.length_median),
                        od(track.width_median),
                        info.time_at(track.first_frame()),
                        info.time_at(track.last_frame()),
                        track.first_frame(),
                        track.last_frame(),
                    ])
                progress_i += 1
                yield f"Generando análisis por hormiga\n{info.video_name}", progress_i, progress_max

            # endregion
            # region Por video
            EN_tracks = [track for track in tracks if EN(track, info)]
            SN_tracks = [track for track in tracks if SN(track, info)]

            progress_i = 0
            progress_max = 16
            data = {'EN': {'l': {}, 'u': {}}, 'SN': {'l': {}, 'u': {}}}
            for k1, tracks in zip(('EN', 'SN'), (EN_tracks, SN_tracks)):
                for k2, load in zip(('l', 'u'), (loaded, unloaded)):
                    for prop, conv in zip(props, (od, area, od, od)):
                        data[k1][k2][prop] = conv(np.mean([
                            getattr(track, prop) for track in load(tracks)
                        ]))
                        progress_i += 1
                        yield f"Generando análisis por video\n{info.video_name}", progress_i, progress_max

            video_sheet.append([
                info.video_name,
                info.start_time,
                info.end_time,
                len(list(loaded(EN_tracks))),
                len(list(unloaded(EN_tracks))),
                len(list(loaded(SN_tracks))),
                len(list(unloaded(SN_tracks))),
                data['EN']['l']['speed_mean'],
                data['EN']['u']['speed_mean'],
                data['SN']['l']['speed_mean'],
                data['SN']['u']['speed_mean'],
                data['EN']['l']['area_median'],
                data['EN']['u']['area_median'],
                data['SN']['l']['area_median'],
                data['SN']['u']['area_median'],
                data['EN']['l']['length_median'],
                data['EN']['u']['length_median'],
                data['SN']['l']['length_median'],
                data['SN']['u']['length_median'],
                data['EN']['l']['width_median'],
                data['EN']['u']['width_median'],
                data['SN']['l']['width_median'],
                data['SN']['u']['width_median'],
            ])
            # endregion

            start_times[trkfile] = info.start_time
            end_times[trkfile] = info.end_time
            global_start_time = info.start_time if itrk == 0 else min(info.start_time, global_start_time)
            global_end_time = info.end_time if itrk == 0 else max(info.end_time, global_end_time)

        lastrow = len(trkfiles) + 2
        video_sheet.append(["Total", f"=MIN(B3:B{lastrow})", f"=MAX(C3:C{lastrow})"] +
                  [f"=SUM({c}3:{c}{lastrow})" for c in "DEFG"] +
                  [f"=AVERAGE({c}3:{c}{lastrow})" for c in "HIJKLMNOPQRSTUVW"])
        video_sheet.cell(lastrow + 1, 2).number_format = 'yyyy-mm-dd h:mm:ss'
        video_sheet.cell(lastrow + 1, 3).number_format = 'yyyy-mm-dd h:mm:ss'

        # TODO: esto podría ir en el loop de arriba. Se necesita primero resolver #15.
        # region En el tiempo
        progress_i = 0
        progress_max = (global_end_time - global_start_time) // time_delta + 1
        yield "Generando análisis en el tiempo", progress_i, progress_max
        time = global_start_time
        while time < global_end_time:
            totals = {'en-load': 0, 'en-ntld': 0, 'sn-load': 0, 'sn-ntld': 0}
            speeds = {'en-load': [], 'en-ntld': [], 'sn-load': [], 'sn-ntld': []}
            areas = {'en-load': [], 'en-ntld': [], 'sn-load': [], 'sn-ntld': []}
            lengths = {'en-load': [], 'en-ntld': [], 'sn-load': [], 'sn-ntld': []}
            widths = {'en-load': [], 'en-ntld': [], 'sn-load': [], 'sn-ntld': []}
            for itrk, trkfile in enumerate(trkfiles):
                start_time = start_times[trkfile]
                end_time = end_times[trkfile]
                #* Si no hay overlap entre el período del `info` y el período de pooling...
                if not (min(time + time_delta, end_time) > max(time, start_time)):
                    continue
                #* Si ya tenemos cargado el info correspondiente a este trkfile...
                # (puede volver a cargar infos innecesariamente si se renombran videos o trkfiles,
                # pero salvo que intercambien nombres entre dos archivos no debería fallar)
                if os.path.splitext(info.video_name)[0] != trkfile.stem:
                    info: TracksCompleteInfo = TracksCompleteInfo.load(trkfile)
                _filter = info.filter_func(**C.TRACKFILTER)
                tracks = [track for track in info.tracks_in_time(time, time + time_delta)
                          if track.load_detected and _filter(track)]
                if not tracks: continue

                od = partial(onedim_conversion, mm_per_pixel=info.mm_per_pixel)
                area = partial(area_conversion, mm_per_pixel=info.mm_per_pixel)

                for track in tracks:
                    direction = info.track_direction(track)
                    if direction == Direction.UN:
                        continue
                    key = (   ('en' if direction == Direction.EN else 'sn') +
                        '-' + ('load' if track.load_prediction else 'ntld'))

                    totals[key] += 1
                    speeds[key].append(od(track.speed_mean))
                    areas[key].append(area(track.area_median))
                    lengths[key].append(od(track.length_median))
                    widths[key].append(od(track.width_median))

            time_sheet.append([
                time,
                time + time_delta,
                totals['en-load'],
                totals['en-ntld'],
                totals['sn-load'],
                totals['sn-ntld'],
                np.mean(speeds['en-load']),
                np.mean(speeds['en-ntld']),
                np.mean(speeds['sn-load']),
                np.mean(speeds['sn-ntld']),
                np.mean(areas['en-load']),
                np.mean(areas['en-ntld']),
                np.mean(areas['sn-load']),
                np.mean(areas['sn-ntld']),
                np.mean(lengths['en-load']),
                np.mean(lengths['en-ntld']),
                np.mean(lengths['sn-load']),
                np.mean(lengths['sn-ntld']),
                np.mean(widths['en-load']),
                np.mean(widths['en-ntld']),
                np.mean(widths['sn-load']),
                np.mean(widths['sn-ntld']),
            ])
            time += time_delta
            progress_i += 1
            yield "Generando análisis en el tiempo", progress_i, progress_max
        # endregion

        for i_ws, ws in enumerate(wb.worksheets):
            adjust_column_widths(ws)
            adjust_decimal_places(ws)
            if i_ws in (0, 2):
                center_headers(ws, 2)
            else:
                center_headers(ws)

        # Los datos se calculan en órden: hormiga-video-tiempo, porque 'hormiga' indica
        # progreso más rápido que las demás y da a entender más rápidamente que hay trabajo
        # en proceso. Sin embargo, el órden de hojas tiene que ser: video-hormiga-tiempo
        wb.move_sheet("Por video", -1)

        self.__wb = wb
        yield "Finalizado", 1, 1
